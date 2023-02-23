# COPYRIGHT ⓒ 2021 HANYANG UNIVERSITY. ALL RIGHTS RESERVED.

import csv
import logging
import numpy as np
import random
from os import makedirs
from datetime import datetime
from route_planner.block import Block
from route_planner.constants import COLOR_LIST
from route_planner.util import log_decorator


class Alloc(object):
    def __init__(self):
        # 할당셀 집합의 목록
        # logging.getLogger('alloc').getLogger(__name__).info('alloc init')
        self.allocated_cells = {}
    
    # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j]에 현재 셀 저장과 새로운 할당셀 집합 생성
    def add_allocate_cell(self, k: int, j: int, allocate_cell: list):
        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j]에 현재 셀 저장과 새로운 할당셀 집합 생성
        self.allocated_cells.setdefault(f'k{k}', {}).setdefault(f'j{j + 1}', []).extend(allocate_cell)
        logging.getLogger('alloc').debug(f'현재 할당셀[D{k}{j+1}]({allocate_cell}을 할당셀 집합에 저장, k증가 k: {k+1}')
        # 현재 할당셀, k 증가
        allocate_cell, k = [], k + 1
        return allocate_cell, k
    
    def sorted_allocated_cells(self):
        _allocated_cells = {k: {k_j: v_j for k_j, v_j in sorted(j.items(), key=lambda y: int(y[0][1:])) if v_j} for k, j in sorted(self.allocated_cells.items(), key=lambda x: int(x[0][1:]))}
        return {k: v for k, v in _allocated_cells.items() if v}

    def reverse_allocated_cells(self, converted_block):
        reversed_allocated_cells = {}
        for k, v in self.allocated_cells.items():
            for j, allocate_cell in v.items():
                if allocate_cell:
                    reversed_allocated_cells.setdefault(k, {})[f'j{converted_block.get(allocate_cell[0]).get("j")}'] = allocate_cell

        self.allocated_cells = reversed_allocated_cells

    # 할당셀 목록 확인
    def check_allocate_cell_j(self, allocate_cell: list, j: int):
        if not allocate_cell:
            for cell in list(self.allocated_cells.values()):
                if f'j{str(j + 1)}' in cell:
                    return True
            return False
        else:
            return True
    
    def convert_allocated_cell(self):
        _converted_allocated_cells = {}
        for k, j_block in self.allocated_cells.items():
            for key_j, values_j in j_block.items():
                for j in values_j:
                    _converted_allocated_cells[j] = {'k': k[1:], 'j': key_j[1:]}
        return _converted_allocated_cells

    # output excel
    @staticmethod
    @log_decorator('할당셀 알고리즘 결과 저장')
    def save_output_csv(output_path: str, input_file_name: str, equipment: str, rearranged_block: list, allocated_cells: dict, converted_block: dict):
        _rearranged_block = rearranged_block.copy()
        _rearranged_block.reverse()
        now = datetime.now().strftime('%Y%m%d%H%M%S')
        makedirs(output_path, exist_ok=True)
        _converted_allocated_cells = {}
        for k, j_block in allocated_cells.items():
            for key_j, values_j in j_block.items():
                for j in values_j:
                    _converted_allocated_cells[j] = {'k': k[1:], 'j': key_j[1:]}
        output = []
        for row, i_block in enumerate(_rearranged_block):
            i_output = []
            for column, block_name in enumerate(i_block):
                k, j = map(lambda x: _converted_allocated_cells.get(block_name, {}).get(x, 'v'), ['k', 'j'])
                dkj = f'[D{k}{j}]' if block_name != Block.VIRTUAL_BLOCK_NAME else '가상셀(이동불가)'
                i_output.append(f'{block_name} {dkj}')
            output.append(i_output)

        with open(f'{output_path}/{input_file_name}_1-2_output.csv', 'w', newline='\n') as csvfile:
            cw = csv.writer(csvfile)
            for i_output in output:
                cw.writerow(i_output)

    # output excel
    @staticmethod
    def save_output_excel(equipment: str, rearranged_block: list, allocated_cells: dict, converted_block: dict):
        import openpyxl as oxl
        from openpyxl.styles.borders import Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, PatternFill

        now = datetime.now().strftime('%Y%m%d%H%M%S')
        _converted_allocated_cells = {}
        for k, j_block in allocated_cells.items():
            for key_j, values_j in j_block.items():
                for j in values_j:
                    _converted_allocated_cells[j] = {'k': k[1:], 'j': key_j[1:]}

        color_map = {'N': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')}
        _rearranged_block = rearranged_block.copy()
        _rearranged_block.reverse()
        wb = oxl.Workbook()
        ws = wb.active
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sample_colors = COLOR_LIST.copy()
        for row, i_block in enumerate(_rearranged_block):
            for column, block_name in enumerate(i_block):
                k, j = map(lambda x: _converted_allocated_cells.get(block_name, {}).get(x, 'v'), ['k', 'j'])
                if not color_map.get(k) and k != 'v':
                    color_number = sample_colors.pop() if sample_colors else ''.join([random.choice('0123456789ABCDEF') for j in range(6)])
                    color_map[k] = PatternFill(start_color=color_number, end_color=color_number, fill_type='solid')
                
                dkj = f'[D{k}{j}]' if block_name != Block.VIRTUAL_BLOCK_NAME else '가상셀(이동불가)'
                ws_cell = ws.cell(row=row + 1,column=column + 1,value=f'{block_name}\n{dkj}')
                ws_cell.border = thin_border
                ws_cell.alignment = Alignment(wrapText=True, horizontal='center')
                ws.column_dimensions[get_column_letter(column + 1)].width = 15
                k = 'N' if converted_block.get(block_name).get('YN') == 'N' and block_name != Block.VIRTUAL_BLOCK_NAME else k
                if color_map.get(k):
                    ws_cell.fill = color_map.get(k)
            ws.row_dimensions[row].height = 40
        
        for c_i, (k, color) in enumerate(sorted(color_map.items(), key=lambda x: int(x[0]) if x[0] != 'N' else 0)):
            ws.cell(len(_rearranged_block) + 2 + c_i, column=1).fill = color
            ws.cell(len(_rearranged_block) + 2 + c_i, column=2, value=f'k={k}' if k !='N' else '이동불가지역')
        wb.save(f'./output/csv/{now}_{equipment}_1-2_output.xlsx')

class Grader(Alloc):
    def __init__(self):
        super().__init__()

    # 할당셀 지정
    # s: 각 셀의 종축길이, l: 룰러 1회 최대 전진거리
    @log_decorator('모터그레이더 할당셀 알고리즘')
    def group(self, rearranged_block: list, converted_block: dict, s: int, l: int):
        # l/s 반올림한 값
        q = round(l / s)
        for j in range(0, max([len(i_block) for i_block in rearranged_block])):
            k, allocate_cell = 1, []
            for i, i_block in enumerate(rearranged_block):
                block_name = i_block[j]
                block = converted_block.get(i_block[j])
                # 가상셀이 아니거나 셀의 이동불가지역이 Y인 경우
                logging.getLogger('alloc').debug(f'현재 셀({block_name})의 이동불가지역 = Y or 현재 셀이 가상셀이 아닌가? {block_name != Block.VIRTUAL_BLOCK_NAME and block.get("YN") == "Y"}')
                if block_name != Block.VIRTUAL_BLOCK_NAME and block.get('YN') == 'Y':
                    # 현재 할당셀 내의 셀 개수가 Q개인가?
                    logging.getLogger('alloc').debug(f'현재 할당셀({allocate_cell}) 내의 셀의 개수({len(allocate_cell)}가 Q({q})개인가? {len(allocate_cell) == q}')
                    if len(allocate_cell) == q:
                        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j]에 현재 셀 저장과 새로운 할당셀 집합 생성
                        # 현재 할당셀, k 증가
                        allocate_cell, k = self.add_allocate_cell(k, j, allocate_cell)
                        allocate_cell.append(block_name)
                    else:
                        # 현재 열에 저장된 할당셀[Dkj] 목록이 있는가?
                        logging.getLogger('alloc').debug(f'현재 열에 저장된 할당셀[Dkj] 목록이 있는가? {self.check_allocate_cell_j(allocate_cell, j)}')
                        if self.check_allocate_cell_j(allocate_cell, j):
                            # 현재 셀과 같은열에 있는 이전 행의 셀이 가상셀인가?
                            logging.getLogger('alloc').debug(f'현재 셀({block_name})과 같은열에 있는 이전 행의 셀이 가상셀인가? {rearranged_block[i - 1][j]} {i > 0 and converted_block.get(rearranged_block[i - 1][j]).get("virtual", False)}')
                            if i > 0 and converted_block.get(rearranged_block[i - 1][j]).get('virtual', False):
                                # 현재 열의 번호 == 1이거나 현재셀과 같은 행에 있는 이전 열의 셀이 가상셀인가?
                                logging.getLogger('alloc').debug(f'현재 열의 번호{j + 1} == 1 {j == 0}')
                                logging.getLogger('alloc').debug(f'현재셀({block_name}) 같은 행에 있는 이전 열의 셀({rearranged_block[i][j - 1]}이 가상셀인가? {converted_block.get(rearranged_block[i][j - 1]).get("virtual", False)}')
                                if j == 0 or converted_block.get(rearranged_block[i][j - 1]).get('virtual', False):
                                    logging.getLogger('alloc').debug(f'k({max((i + 1) // q, k)}) = Max(현재 행의 번호 / Q 의 몫({(i + 1) // q}), 현재 k({k}))')
                                    k = max((i + 1) // q, k)
                                else:
                                    # j-1열의 k를 사용
                                    for a_k, a_v in self.allocated_cells.items():
                                        for j_a_k, j_a_v in a_v.items():
                                            if rearranged_block[i][j - 1] in j_a_v:
                                                k = int(a_k[1:])
                                                break
                                    logging.getLogger('alloc').debug(f'k = 같은행의 j-1열에 있는 셀의 k, k: {k}')
                        else:
                            # 현재 행의 번호 / Q 가 자연수인가?
                            logging.getLogger('alloc').debug(f'현재 행의 번호({i+1}) / Q({q}) 가 자연수인가? {(i + 1) / q - int((i + 1) / q) == 0}')
                            if (i + 1) / q - int((i + 1) / q) == 0:
                                # k = Max{(현재 행의 번호 / Q) 의 몫, 현재 k}
                                logging.getLogger('alloc').debug(f'k({max((i + 1) // q, k)}) = Max(현재 행의 번호({i + 1}) / Q({q}) 의 몫({(i + 1) // q}), 현재 k({k}))')
                                k = max((i + 1) // q, k)
                            else:
                                # k = Max{(현재 행의 번호 / Q) 의 몫 + 1, 현재 k}
                                logging.getLogger('alloc').debug(f'k({max(((i + 1) // q) + 1, k)}) = Max(현재 행의 번호({i + 1}) / Q({q}) 의 몫 + 1({((i + 1) // q) + 1}), 현재 k({k}))')
                                k = max(((i + 1) // q) + 1, k)
                        
                        allocate_cell.append(block_name)
                else:
                    if allocate_cell:
                        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j] 집합 생성
                        # 현재 할당셀, k 증가
                        allocate_cell, k = self.add_allocate_cell(k, j, allocate_cell)
                # 현재 행의 번호가 마지막일 경우, 현재 할당셀[Dkj]을 할당셀 집합에 저장
                if i == len(rearranged_block) - 1:
                    allocate_cell, _ = self.add_allocate_cell(k, j, allocate_cell)

        return self.sorted_allocated_cells()

class Paver(Alloc):
    def __init__(self):
        super().__init__()

    # 할당셀 지정
    # s: 각 셀의 종축길이, l: 룰러 1회 최대 전진거리
    @log_decorator('페이버 할당셀 알고리즘')
    def group(self, rearranged_block: list, converted_block: dict, s: int, l: int):
        # l/s 반올림한 값
        q = round(l / s)
        for j in range(0, max([len(i_block) for i_block in rearranged_block])):
            k, allocate_cell = 1, []
            for i, i_block in enumerate(rearranged_block):
                block_name = i_block[j]
                block = converted_block.get(i_block[j])
                # 가상셀이 아니거나 셀의 이동불가지역이 Y인 경우
                logging.getLogger('alloc').debug(f'현재 셀({block_name})의 이동불가지역 = Y or 현재 셀이 가상셀이 아닌가? {block_name != Block.VIRTUAL_BLOCK_NAME and block.get("YN") == "Y"}')
                if block_name != Block.VIRTUAL_BLOCK_NAME and block.get('YN') == 'Y':
                    # 현재 할당셀 내의 셀 개수가 Q개인가?
                    logging.getLogger('alloc').debug(f'현재 할당셀({allocate_cell}) 내의 셀의 개수({len(allocate_cell)}가 Q({q})개인가? {len(allocate_cell) == q}')
                    if len(allocate_cell) == q:
                        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j]에 현재 셀 저장과 새로운 할당셀 집합 생성
                        # 현재 할당셀, k 증가
                        allocate_cell, k = self.add_allocate_cell(k, j, allocate_cell)
                        allocate_cell.append(block_name)
                    else:
                        # 현재 열에 저장된 할당셀[Dkj] 목록이 있는가?
                        logging.getLogger('alloc').debug(f'현재 열에 저장된 할당셀[Dkj] 목록이 있는가? {self.check_allocate_cell_j(allocate_cell, j)}')
                        if self.check_allocate_cell_j(allocate_cell, j):
                            # 현재 셀과 같은열에 있는 이전 행의 셀이 가상셀인가?
                            logging.getLogger('alloc').debug(f'현재 셀({block_name})과 같은열에 있는 이전 행의 셀이 가상셀인가? {rearranged_block[i - 1][j]} {i > 0 and converted_block.get(rearranged_block[i - 1][j]).get("virtual", False)}')
                            if i > 0 and converted_block.get(rearranged_block[i - 1][j]).get('virtual', False):
                                # 현재 열의 번호 == 1이거나 현재셀과 같은 행에 있는 이전 열의 셀이 가상셀인가?
                                logging.getLogger('alloc').debug(f'현재 열의 번호{j + 1} == 1 {j == 0}')
                                logging.getLogger('alloc').debug(f'현재셀({block_name}) 같은 행에 있는 이전 열의 셀({rearranged_block[i][j - 1]}이 가상셀인가? {converted_block.get(rearranged_block[i][j - 1]).get("virtual", False)}')
                                if j == 0 or converted_block.get(rearranged_block[i][j - 1]).get('virtual', False):
                                    logging.getLogger('alloc').debug(f'k({max((i + 1) // q, k)}) = Max(현재 행의 번호 / Q 의 몫({(i + 1) // q}), 현재 k({k}))')
                                    k = max((i + 1) // q, k)
                                else:
                                    # j-1열의 k를 사용
                                    for a_k, a_v in self.allocated_cells.items():
                                        for j_a_k, j_a_v in a_v.items():
                                            if rearranged_block[i][j - 1] in j_a_v:
                                                k = int(a_k[1:])
                                                break
                                    logging.getLogger('alloc').debug(f'k = 같은행의 j-1열에 있는 셀의 k, k: {k}')
                        else:
                            # 현재 행의 번호 / Q 가 자연수인가?
                            logging.getLogger('alloc').debug(f'현재 행의 번호({i+1}) / Q({q}) 가 자연수인가? {(i + 1) / q - int((i + 1) / q) == 0}')
                            if (i + 1) / q - int((i + 1) / q) == 0:
                                # k = Max{(현재 행의 번호 / Q) 의 몫, 현재 k}
                                logging.getLogger('alloc').debug(f'k({max((i + 1) // q, k)}) = Max(현재 행의 번호({i + 1}) / Q({q}) 의 몫({(i + 1) // q}), 현재 k({k}))')
                                k = max((i + 1) // q, k)
                            else:
                                # k = Max{(현재 행의 번호 / Q) 의 몫 + 1, 현재 k}
                                logging.getLogger('alloc').debug(f'k({max(((i + 1) // q) + 1, k)}) = Max(현재 행의 번호({i + 1}) / Q({q}) 의 몫 + 1({((i + 1) // q) + 1}), 현재 k({k}))')
                                k = max(((i + 1) // q) + 1, k)
                        
                        allocate_cell.append(block_name)
                else:
                    if allocate_cell:
                        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j] 집합 생성
                        # 현재 할당셀, k 증가
                        allocate_cell, k = self.add_allocate_cell(k, j, allocate_cell)
                # 현재 행의 번호가 마지막일 경우, 현재 할당셀[Dkj]을 할당셀 집합에 저장
                if i == len(rearranged_block) - 1:
                    allocate_cell, _ = self.add_allocate_cell(k, j, allocate_cell)

        return self.sorted_allocated_cells()

class Roller(Alloc):
    def __init__(self):
        super().__init__()

    # 할당셀 지정
    # s: 각 셀의 종축길이, l: 룰러 1회 최대 전진거리
    @log_decorator('롤러 할당셀 알고리즘')
    def group(self, rearranged_block: list, converted_block: dict, s: int, l: int):
        # l/s 반올림한 값
        q = round(l / s)
        for j in range(0, max([len(i_block) for i_block in rearranged_block])):
            k, allocate_cell = 1, []
            for i, i_block in enumerate(rearranged_block):
                block_name = i_block[j]
                block = converted_block.get(i_block[j])
                # 가상셀이 아니거나 셀의 이동불가지역이 Y인 경우
                logging.getLogger('alloc').debug(f'현재 셀({block_name})의 이동불가지역 = Y or 현재 셀이 가상셀이 아닌가? {block_name != Block.VIRTUAL_BLOCK_NAME and block.get("YN") == "Y"}')
                if block_name != Block.VIRTUAL_BLOCK_NAME and block.get('YN') == 'Y':
                    # 현재 할당셀 내의 셀 개수가 Q개인가?
                    logging.getLogger('alloc').debug(f'현재 할당셀({allocate_cell}) 내의 셀의 개수({len(allocate_cell)}가 Q({q})개인가? {len(allocate_cell) == q}')
                    if len(allocate_cell) == q:
                        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j]에 현재 셀 저장과 새로운 할당셀 집합 생성
                        # 현재 할당셀, k 증가
                        allocate_cell, k = self.add_allocate_cell(k, j, allocate_cell)
                        allocate_cell.append(block_name)
                    else:
                        # 현재 열에 저장된 할당셀[Dkj] 목록이 있는가?
                        logging.getLogger('alloc').debug(f'현재 열에 저장된 할당셀[Dkj] 목록이 있는가? {self.check_allocate_cell_j(allocate_cell, j)}')
                        if self.check_allocate_cell_j(allocate_cell, j):
                            # 현재 셀과 같은열에 있는 이전 행의 셀이 가상셀인가?
                            logging.getLogger('alloc').debug(f'현재 셀({block_name})과 같은열에 있는 이전 행의 셀이 가상셀인가? {rearranged_block[i - 1][j]} {i > 0 and converted_block.get(rearranged_block[i - 1][j]).get("virtual", False)}')
                            if i > 0 and converted_block.get(rearranged_block[i - 1][j]).get('virtual', False):
                                # 현재 열의 번호 == 1이거나 현재셀과 같은 행에 있는 이전 열의 셀이 가상셀인가?
                                logging.getLogger('alloc').debug(f'현재 열의 번호{j + 1} == 1 {j == 0}')
                                logging.getLogger('alloc').debug(f'현재셀({block_name}) 같은 행에 있는 이전 열의 셀({rearranged_block[i][j - 1]}이 가상셀인가? {converted_block.get(rearranged_block[i][j - 1]).get("virtual", False)}')
                                if j == 0 or converted_block.get(rearranged_block[i][j - 1]).get('virtual', False):
                                    logging.getLogger('alloc').debug(f'k({max((i + 1) // q, k)}) = Max(현재 행의 번호 / Q 의 몫({(i + 1) // q}), 현재 k({k}))')
                                    k = max((i + 1) // q, k)
                                else:
                                    # j-1열의 k를 사용
                                    for a_k, a_v in self.allocated_cells.items():
                                        for j_a_k, j_a_v in a_v.items():
                                            if rearranged_block[i][j - 1] in j_a_v:
                                                k = int(a_k[1:])
                                                break
                                    logging.getLogger('alloc').debug(f'k = 같은행의 j-1열에 있는 셀의 k, k: {k}')
                        else:
                            # 현재 행의 번호 / Q 가 자연수인가?
                            logging.getLogger('alloc').debug(f'현재 행의 번호({i+1}) / Q({q}) 가 자연수인가? {(i + 1) / q - int((i + 1) / q) == 0}')
                            if (i + 1) / q - int((i + 1) / q) == 0:
                                # k = Max{(현재 행의 번호 / Q) 의 몫, 현재 k}
                                logging.getLogger('alloc').debug(f'k({max((i + 1) // q, k)}) = Max(현재 행의 번호({i + 1}) / Q({q}) 의 몫({(i + 1) // q}), 현재 k({k}))')
                                k = max((i + 1) // q, k)
                            else:
                                # k = Max{(현재 행의 번호 / Q) 의 몫 + 1, 현재 k}
                                logging.getLogger('alloc').debug(f'k({max(((i + 1) // q) + 1, k)}) = Max(현재 행의 번호({i + 1}) / Q({q}) 의 몫 + 1({((i + 1) // q) + 1}), 현재 k({k}))')
                                k = max(((i + 1) // q) + 1, k)
                        
                        allocate_cell.append(block_name)
                else:
                    if allocate_cell:
                        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j] 집합 생성
                        # 현재 할당셀, k 증가
                        allocate_cell, k = self.add_allocate_cell(k, j, allocate_cell)
                # 현재 행의 번호가 마지막일 경우, 현재 할당셀[Dkj]을 할당셀 집합에 저장
                if i == len(rearranged_block) - 1:
                    allocate_cell, _ = self.add_allocate_cell(k, j, allocate_cell)

        return self.sorted_allocated_cells()

class Dozer(Alloc):
    def __init__(self):
        super().__init__()
        self.converted_allocated_cells = {}

    @log_decorator('도저 절토 할당셀 알고리즘')
    def cut_group(self, rearranged_block: list, converted_block: dict, e: int, s: int, l: int, allowable_error_height: float):
        return self.group(Block.reverse_block(rearranged_block), converted_block, e, s, l, allowable_error_height)

    @log_decorator('도저 성토 할당셀 알고리즘')
    def fill_group(self, rearranged_block: list, converted_block: dict, e: int, s: int, l: int, allowable_error_height: float):
        return self.group(rearranged_block, converted_block, e, s, l, allowable_error_height)

    # 할당셀 지정
    # e: 불도저 작업역량, s: 각 셀의 종축길이, l: 종축 1회 최대 정지거리, allowable_error_height: 허용 오차 높이
    def group(self, rearranged_block: list, converted_block: dict, e: int, s: int, l: int, allowable_error_height: float):
        # l/s 반올림한 값
        q = round(l / s)
        earth_volume, temp_cells = 0, []
        epsilon = max([len(i_block) for i_block in rearranged_block])
        for j in range(0, epsilon):
            k, allocate_cell, earth_volume = 1, [], 0
            for i, i_block in enumerate(rearranged_block):
                block_name = i_block[j]
                block = converted_block.get(i_block[j])

                # 가상셀이 아니거나 셀의 이동불가지역이 Y인 경우
                logging.getLogger('alloc').debug(f'현재 셀({block_name})의 이동불가지역 = Y or 현재 셀이 가상셀이 아닌가? {block_name != Block.VIRTUAL_BLOCK_NAME and block.get("YN") == "Y"}')
                if block_name != Block.VIRTUAL_BLOCK_NAME and block.get('YN') == 'Y':
                    # cutVol: 절토, fillVol: 성토, Area: 면적
                    cut_vol, fill_vol, area = map(lambda x: float(block.get(x)), ['cutVol', 'fillVol', 'Area'])
                    # 현재 셀의 (성토-절토) < 현재 셀의 면적 * 허용 오차 높이
                    logging.getLogger('alloc').debug(f'현재 셀({block_name})의 성토({fill_vol})-절토({cut_vol}) < 현재 셀의 면적({area})*허용오차높이({allowable_error_height}) {fill_vol - cut_vol < area * allowable_error_height}')
                    if fill_vol - cut_vol < area * allowable_error_height:
                        allocate_cell, earth_volume, k, temp_cells = self.add_dozer_allocate_cell(k, j, allocate_cell, earth_volume, block, temp_cells)
                    else:
                        # 현재 할당셀 내의 셀의 개수가 Q개인가?
                        # or 현재 셀의 (성토 - 절토) + 현재 누적토공량  <= e (작업역량)이 아닐 경우
                        logging.getLogger('alloc').debug(f'현재 할당셀({allocate_cell}) 내의 셀의 개수({len(allocate_cell)}가 Q({q})개인가? {len(allocate_cell) == q}')
                        logging.getLogger('alloc').debug(f'현재 셀({block_name})의 성토({fill_vol})-절토({cut_vol})+현재누적토공량({earth_volume}) <= {e} {(fill_vol - cut_vol + earth_volume <= e)}')
                        if len(allocate_cell) == q:
                            # 임시변수에 셀을 저장
                            temp_cells.append(block)
                            allocate_cell, earth_volume, k, temp_cells = self.add_dozer_allocate_cell(k, j, allocate_cell, earth_volume, block, temp_cells)
                        elif not (fill_vol - cut_vol + earth_volume <= e):
                            if len(allocate_cell) == 0:
                                # 현재 셀을 현재 할당셀 Dkj에 저장
                                allocate_cell.append(block.get('BlName'))
                                logging.getLogger('alloc').debug(f'현재셀({block.get("BlName")})을 현재 할당셀에 저장, 할당셀: {allocate_cell}')
                            else:
                                #임시 변수에 셀을 저장
                                temp_cells.append(block)
                                allocate_cell, earth_volume, k, temp_cells = self.add_dozer_allocate_cell(k, j, allocate_cell, earth_volume, block, temp_cells)
                        else:
                            # 현재 열에 저장된 할당셀[Dkj] 목록이 없는가?
                            logging.getLogger('alloc').debug(f'현재 열에 저장된 할당셀[Dkj] 목록이 없는가? {not self.check_allocate_cell_j(allocate_cell, j)}')
                            if not self.check_allocate_cell_j(allocate_cell, j):
                                # 현재 행의 번호 /Q 의 계산 결과 중 나머지의 값이 0인가?
                                logging.getLogger('alloc').debug(f'현재 행의 번호({i+1}) /Q({q}) 의 계산 결과 중 나머지의 값({(i + 1) % q})이 0인가? {(i + 1) % q == 0}')
                                k = (i + 1) // q if (i + 1) % q == 0 else ((i + 1) // q) + 1
                                logging.getLogger('alloc').debug(f'k: {k}')
                            else:
                                # 현재 셀과 같은열에 있는 이전 행의 셀이 가상셀인가?
                                logging.getLogger('alloc').debug(f'현재 셀과 같은열에 있는 이전 행의 셀이 가상셀인가? {i > 0 and converted_block.get(rearranged_block[i - 1][j]).get("virtual", False)}')
                                if i > 0 and converted_block.get(rearranged_block[i - 1][j]).get('virtual', False):
                                    # 현재 셀과 같은행에 있는 이전 열의 셀이 가상셀인가?
                                    logging.getLogger('alloc').debug(f'현재 셀과 같은행에 있는 이전 행의 셀이 가상셀인가? {j > 0 and converted_block.get(rearranged_block[i][j - 1]).get("virtual", False)}')
                                    if j > 0 and converted_block.get(rearranged_block[i][j - 1]).get("virtual", False):
                                        # max_bef_i: ("'행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀'의 행번호") 중 MAX값
                                        max_bef_i = max([__i + 1 for __i in range(0, len(rearranged_block))  if not converted_block.get(rearranged_block[__i][j]).get("virtual", False) and __i < i])
                                        if not self.converted_allocated_cells.get(rearranged_block[max_bef_i - 1][j]):
                                            self.converted_allocated_cells = self.convert_allocated_cell()
                                        # max_bef_k: ("'행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀'의 행번호") 중 MAX값에 해당하는 셀의 k
                                        max_bef_k = int(self.converted_allocated_cells.get(rearranged_block[max_bef_i - 1][j]).get('k'))

                                        # (현재 셀의 행번호 - (행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀의 행번호 중 MAX 값) - 1) / Q의 나머지가 0 인가?
                                        logging.getLogger('alloc').debug(f'({i + 1}(현재 셀의 행번호) - {max_bef_i}(행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀의 행번호{[__i + 1 for __i in range(0, len(rearranged_block))  if not converted_block.get(rearranged_block[__i][j]).get("virtual", False) and __i < i]} 중 MAX 값) - 1) / {q}(Q)의 나머지가 0 인가? {((i + 1) - max_bef_i - 1) % q == 0}')
                                        if ((i + 1) - max_bef_i - 1) % q == 0:
                                            # k = 행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀 중 행번호가 제일 큰 셀의 k값 + ((현재 셀의 행번호 - (행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀의 행번호 중 MAX 값) - 1) / Q 의 몫)
                                            k = max_bef_k + (((i + 1) - max_bef_i - 1) // q)
                                            logging.getLogger('alloc').debug(f'{k}(k) = {max_bef_k}(행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀 중 행번호가 제일 큰 셀의 k값) + {(((i + 1) - max_bef_i - 1) // q)}(({i + 1}(현재 셀의 행번호) - {max_bef_i}(행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀의 행번호 중 MAX 값) - 1) / {q}(Q) 의 몫)')
                                        else:
                                            # k = 행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀 중 행번호가 제일 큰 셀의 k값 + ((현재 셀의 행번호 - (행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀의 행번호 중 MAX 값) - 1) / Q 의 몫) + 1
                                            k = max_bef_k + (((i + 1) - max_bef_i - 1) // q) + 1
                                            logging.getLogger('alloc').debug(f'{k}(k) = {max_bef_k}(행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀 중 행번호가 제일 큰 셀의 k값) + {(((i + 1) - max_bef_i - 1) // q)}(({i + 1}(현재 셀의 행번호) - {max_bef_i}(행번호 < 현재 셀의 행번호 & 가상셀이 아닌 셀의 행번호 중 MAX 값) - 1) / {q}(Q) 의 몫) + 1')
                                    else:
                                        # j-1열의 k를 사용
                                        if j > 0:
                                            for a_k, a_v in self.allocated_cells.items():
                                                for j_a_k, j_a_v in a_v.items():
                                                    if rearranged_block[i][j - 1] in j_a_v:
                                                        k = int(a_k[1:])
                                                        break
                                            logging.getLogger('alloc').debug(f'k = 같은행의 j-1열에 있는 셀의 k, k: {k}')
                                        else:
                                            logging.getLogger('alloc').warning(f'j가 1이라서 j-1열에 있는 셀이 없음, k: {k}, block: {block_name}')

                            #현재 누적토공량에 현재 셀의 (성토-절토)를 더함
                            earth_volume += (float(block.get('fillVol')) - float(block.get('cutVol')))
                            logging.getLogger('alloc').debug(f'현재 누적토공량에 현재셀({block.get("BlName")})의 성토({block.get("fillVol")})-절토({block.get("cutVol")})를 더함')
                            # 현재 셀을 현재 할당셀 Dkj에 저장
                            allocate_cell.append(block.get('BlName'))
                            logging.getLogger('alloc').debug(f'현재셀({block.get("BlName")})을 현재 할당셀에 저장, 할당셀: {allocate_cell}')
                else:
                    logging.getLogger('alloc').debug(f'현재 할당셀 목록에 저장된 셀이 없는가? {len(allocate_cell)>0}')
                    if allocate_cell:
                        allocate_cell, earth_volume, k, temp_cells, = self.add_dozer_allocate_cell(k, j, allocate_cell, earth_volume, block, temp_cells)
                if len(rearranged_block) - 1 == i:
                    # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j]에 현재 셀 저장과 새로운 할당셀 집합 생성
                    logging.getLogger('alloc').debug(f'현재 할당셀[D{k}{j+1}]{allocate_cell}을 할당셀 집합에 저장')
                    self.allocated_cells.setdefault(f'k{k}', {}).setdefault(f'j{j + 1}', []).extend(allocate_cell)
                    logging.getLogger('alloc').debug(f'전체 할당셀 집합을 저장 후 반환')
        self.reverse_allocated_cells(converted_block)
        return self.sorted_allocated_cells()

    # 할당셀 추가
    def add_dozer_allocate_cell(self, k: int, j: int, allocate_cell: list, earth_volume: int, block: dict, temp_cells: list):
        # 현재 할당셀[Dkj]을 할당셀 집합에 저장 후 새로운 할당셀[D[(k+1)j]에 현재 셀 저장과 새로운 할당셀 집합 생성
        # 현재 할당셀, k 증가
        allocate_cell, k = self.add_allocate_cell(k, j, allocate_cell)
        # 누적 토공량 초기화
        logging.getLogger('alloc').debug(f'누적 토공량 초기화({earth_volume}->0), 임시변수에 값이 있는가? {len(temp_cells)>0}')
        earth_volume = 0
        # 임시 변수에 값이 있는가?
        if temp_cells:
            allocate_cell.extend([cell.get('BlName') for cell in temp_cells])
            # 누적 토공량에 임시 변수에 저장된 셀의 토량(성토 - 절토)을 더한 후 임시 변수 초기화
            earth_volume += sum([float(cell.get('fillVol')) - float(cell.get('cutVol')) for cell in temp_cells])
            temp_cells = []
            logging.getLogger('alloc').debug(f'임시 변수에 저장된 셀을 새로운 할당셀에 추가, {allocate_cell}')
            logging.getLogger('alloc').debug(f'D(k+1)j의 누적토공량에 임시 변수에 저장된 셀의 토량을 더한 후 임시 변수 초기화, 누적토공량: {earth_volume}')
        return allocate_cell, earth_volume, k, temp_cells
